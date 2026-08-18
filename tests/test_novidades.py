"""
test_novidades.py
-----------------
Testa o que mudou na versao 1.0:

  1. Login por CATEGORIA (qualquer e-mail + a senha da categoria)
  2. Banco criado automaticamente ao ligar
  3. Envio da planilha .xlsx pela tela
  4. API de integracao (/api/v1)
  5. Nenhuma senha ou aviso de prototipo aparecendo nas telas

COMO RODAR (com o venv ativado, na raiz do projeto):
    python tests/test_novidades.py
"""

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from app import config  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.main import app  # noqa: E402
from app.models import LoginHistory, Payment  # noqa: E402

passou = 0
falhou = 0


def verificar(descricao: str, condicao: bool, detalhe: str = "") -> None:
    global passou, falhou
    if condicao:
        passou += 1
        print(f"  [OK]     {descricao}")
    else:
        falhou += 1
        print(f"  [FALHOU] {descricao}")
        if detalhe:
            print(f"           -> {detalhe}")


def entrar(cliente, email, senha, perfil):
    return cliente.post(
        "/login",
        data={"email": email, "senha": senha, "perfil": perfil},
        follow_redirects=False,
    )


def logar(perfil="ESTIPULANTE", email="teste@qualquer.com"):
    senhas = {
        "ESTIPULANTE": config.SENHA_ESTIPULANTE,
        "CORRETORA": config.SENHA_CORRETORA,
        "SEGURADORA": config.SENHA_SEGURADORA,
    }
    c = TestClient(app)
    entrar(c, email, senhas[perfil], perfil)
    return c


print("\n=== TESTANDO AS NOVIDADES DA VERSAO 1.0 ===\n")

# ---------------------------------------------------------------
# 1. LOGIN POR CATEGORIA
# ---------------------------------------------------------------
print("1. Login por categoria (e-mail livre)")

# Um e-mail que NAO existe no banco deve entrar, com a senha certa.
c = TestClient(app)
r = entrar(c, "pessoa.qualquer@gmail.com", config.SENHA_ESTIPULANTE, "ESTIPULANTE")
verificar("e-mail nunca visto entra com a senha do estipulante",
          r.status_code == 303 and r.headers.get("location") == "/dashboard",
          f"veio {r.status_code}")

pagina = c.get("/dashboard").text
verificar("  a tela mostra o e-mail de quem entrou",
          "pessoa.qualquer@gmail.com" in pagina)
verificar("  e mostra a categoria", "ESTIPULANTE" in pagina)
verificar("  NAO aparece mais 'Luciana Ferraz'", "Luciana Ferraz" not in pagina)

# A senha de uma categoria nao pode servir para outra.
r = entrar(TestClient(app), "alguem@x.com", config.SENHA_CORRETORA, "ESTIPULANTE")
verificar("senha da corretora NAO entra como estipulante",
          r.status_code == 401, f"veio {r.status_code}")

r = entrar(TestClient(app), "alguem@x.com", config.SENHA_SEGURADORA, "CORRETORA")
verificar("senha da seguradora NAO entra como corretora", r.status_code == 401)

# Cada senha entra na sua categoria.
for perfil, senha in [
    ("ESTIPULANTE", config.SENHA_ESTIPULANTE),
    ("CORRETORA", config.SENHA_CORRETORA),
    ("SEGURADORA", config.SENHA_SEGURADORA),
]:
    r = entrar(TestClient(app), "outro@email.com", senha, perfil)
    verificar(f"  senha de {perfil} entra em {perfil}", r.status_code == 303)

r = entrar(TestClient(app), "alguem@x.com", "chute", "ESTIPULANTE")
verificar("senha errada continua sendo recusada", r.status_code == 401)

# O e-mail digitado tem que ficar registrado.
db = SessionLocal()
ultimo = db.query(LoginHistory).order_by(LoginHistory.id.desc()).first()
db.close()
verificar("a tentativa foi registrada com o e-mail digitado",
          ultimo is not None and ultimo.email_informado == "alguem@x.com",
          f"registrado: {ultimo.email_informado if ultimo else '?'}")

# ---------------------------------------------------------------
# 2. NENHUMA SENHA OU AVISO DE PROTOTIPO NAS TELAS
# ---------------------------------------------------------------
print("\n2. Telas sem senhas expostas e sem rotulo de prototipo")

tela_login = TestClient(app).get("/login").text
verificar("a tela de login NAO mostra os e-mails de acesso",
          "estipulante@sebraeprev.com.br" not in tela_login)
verificar("  NAO mostra nenhuma senha",
          config.SENHA_ESTIPULANTE not in tela_login)
verificar("  NAO diz 'ambiente de desenvolvimento'",
          "desenvolvimento" not in tela_login.lower())
verificar("  avisa que o acesso e restrito e registrado",
          "restrito" in tela_login.lower())

cliente = logar()
PALAVRAS_PROIBIDAS = ["protótipo", "prototipo", "dados fictícios", "em desenvolvimento"]
for url in ["/dashboard", "/seguros", "/movimentacao", "/assistente",
            "/produtos", "/integracoes", "/comissoes", "/sinistros"]:
    texto = cliente.get(url).text.lower()
    achadas = [p for p in PALAVRAS_PROIBIDAS if p in texto]
    verificar(f"{url} sem rótulo de protótipo", not achadas, f"achei: {achadas}")

# ---------------------------------------------------------------
# 3. ENVIO DA PLANILHA
# ---------------------------------------------------------------
print("\n3. Envio da planilha .xlsx")


def montar_planilha(linhas, competencia="09/2026", com_situacao=True):
    """Monta um .xlsx em memoria, igual ao que a corretora enviaria."""
    livro = Workbook()
    aba = livro.active
    cabecalho = [
        "NOME DO SEGURADO", "NUMERO DA MATRICULA", "NUMERO DO CPF",
        "VALOR DO CAPITAL MORTE", "VALOR DO CAPITAL INVALIDEZ",
        "VALOR DO PREMIO TOTAL/LIQUIDO", "CODIGO DO MODULO", "CODIGO SUB",
        "COMPETENCIA (dados pagamento)",
    ]
    if com_situacao:
        cabecalho.append("SITUACAO DO PAGAMENTO")
    aba.append(cabecalho)

    for linha in linhas:
        aba.append(list(linha) + [competencia] + ([linha[-1]] if False else []))

    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()


# Planilha valida com 3 segurados, competencia nova 09/2026.
livro = Workbook()
aba = livro.active
aba.append(["NOME DO SEGURADO", "NUMERO DA MATRICULA", "NUMERO DO CPF",
            "VALOR DO CAPITAL MORTE", "VALOR DO CAPITAL INVALIDEZ",
            "VALOR DO PREMIO TOTAL/LIQUIDO", "CODIGO DO MODULO",
            "CODIGO SUB", "COMPETENCIA (dados pagamento)",
            "SITUACAO DO PAGAMENTO"])
aba.append(["Mariana Teste Silva", 900001, "111.222.333-44", 150000, 150000, 60.70, 101, "01", "09/2026", "Pago"])
aba.append(["Joao Teste Souza", 900002, "555.666.777-88", 200000, 200000, "81,10", 101, "01", "09/2026", "A pagar"])
aba.append(["Ana Teste Lima", 900003, "999.888.777-66", 120000, 120000, 50.10, 101, "02", "09/2026", "Em atraso"])
aba.append(["TOTAL", "", "", 470000, 470000, 191.90, "", "", "", ""])
buffer = io.BytesIO()
livro.save(buffer)
planilha_ok = buffer.getvalue()

r = cliente.post("/movimentacao/enviar",
                 files={"arquivo": ("base_09_2026.xlsx", planilha_ok,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
verificar("a planilha foi aceita", r.status_code == 200, f"veio {r.status_code}")
verificar("  a tela confirma o sucesso", "processada com sucesso" in r.text)
verificar("  informa 3 registros gravados", "3 registro" in r.text)
verificar("  a linha TOTAL do rodape foi ignorada", "900004" not in r.text)

db = SessionLocal()
novos = db.query(Payment).filter(Payment.competencia == "09/2026").all()
verificar("  3 registros entraram no banco", len(novos) == 3, f"entraram: {len(novos)}")
verificar("  o premio com virgula foi lido certo (81,10 -> 81.1)",
          any(abs(p.premio - 81.10) < 0.01 for p in novos))
verificar("  a situacao 'Em atraso' foi respeitada",
          any(p.status == "Em atraso" for p in novos))
verificar("  a competencia 07/2026 continua intacta",
          db.query(Payment).filter(Payment.competencia == "07/2026").count() == 10)
db.close()

# Reenviar a MESMA planilha nao pode duplicar.
cliente.post("/movimentacao/enviar",
             files={"arquivo": ("base_09_2026.xlsx", planilha_ok, "application/octet-stream")})
db = SessionLocal()
verificar("reenviar a mesma planilha NAO duplica",
          db.query(Payment).filter(Payment.competencia == "09/2026").count() == 3)
db.close()

# A tela deixa escolher a competencia.
r = cliente.get("/movimentacao?competencia=09/2026")
verificar("da para escolher a competencia na tela",
          r.status_code == 200 and "Mariana Teste Silva" in r.text)

# Arquivo que nao e .xlsx.
r = cliente.post("/movimentacao/enviar",
                 files={"arquivo": ("foto.png", b"nao sou uma planilha", "image/png")})
verificar("arquivo que nao e .xlsx e recusado", "não é um arquivo .xlsx" in r.text)

# Planilha sem as colunas obrigatorias.
livro = Workbook()
livro.active.append(["COLUNA ERRADA", "OUTRA COISA"])
livro.active.append(["a", "b"])
buffer = io.BytesIO()
livro.save(buffer)
r = cliente.post("/movimentacao/enviar",
                 files={"arquivo": ("errada.xlsx", buffer.getvalue(), "application/octet-stream")})
verificar("planilha sem as colunas certas e recusada",
          "colunas obrigatórias" in r.text)

db = SessionLocal()
verificar("  e o banco nao foi tocado pela planilha invalida",
          db.query(Payment).filter(Payment.competencia == "09/2026").count() == 3)
db.close()

# A corretora tambem pode enviar; a seguradora tambem (nao ha bloqueio nesse modulo).
r = logar("CORRETORA").post("/movimentacao/enviar",
                            files={"arquivo": ("x.xlsx", planilha_ok, "application/octet-stream")})
verificar("a corretora tambem consegue enviar planilha", r.status_code == 200)

# ---------------------------------------------------------------
# 4. API DE INTEGRACAO
# ---------------------------------------------------------------
print("\n4. API de integracao (/api/v1)")

api = TestClient(app)
CHAVE = {"X-API-Key": config.API_KEY}

verificar("a API_KEY esta configurada", bool(config.API_KEY),
          "defina API_KEY no .env")

# Sem chave -> 401
r = api.get("/api/v1/status")
verificar("sem a chave, a API responde 401", r.status_code == 401, f"veio {r.status_code}")

# Chave errada -> 401
r = api.get("/api/v1/status", headers={"X-API-Key": "chave-errada"})
verificar("com chave errada, tambem 401", r.status_code == 401)

# Com a chave certa
r = api.get("/api/v1/status", headers=CHAVE)
verificar("com a chave certa, responde 200", r.status_code == 200)
if r.status_code == 200:
    dados = r.json()
    verificar("  informa que esta no ar", dados.get("situacao") == "no ar")
    verificar("  traz o total de apolices", dados["totais"]["apolices"] == 50)

r = api.get("/api/v1/apolices?status=Ativa&limite=5", headers=CHAVE)
verificar("lista apolices filtrando por status",
          r.status_code == 200 and len(r.json()) == 5, f"veio {r.status_code}")
verificar("  todas vieram como Ativa",
          all(a["status"] == "Ativa" for a in r.json()))

r = api.get("/api/v1/apolices?vencendo_em=30", headers=CHAVE)
verificar("filtra apolices que vencem em 30 dias", r.status_code == 200)

r = api.get("/api/v1/apolices/AP-2041", headers=CHAVE)
verificar("busca uma apolice pelo numero",
          r.status_code == 200 and r.json()["participante"] == "Marcos A. Ribeiro")

r = api.get("/api/v1/apolices/NAO-EXISTE", headers=CHAVE)
verificar("apolice inexistente devolve 404", r.status_code == 404)

r = api.get("/api/v1/indicadores", headers=CHAVE)
verificar("traz os indicadores da carteira",
          r.status_code == 200 and "capital_segurado" in r.json())

for rota in ["/api/v1/sinistros", "/api/v1/comissoes", "/api/v1/inadimplencia"]:
    verificar(f"{rota} responde 200",
              api.get(rota, headers=CHAVE).status_code == 200)

# --- POST: receber movimentacao de outro sistema ---
lote = {
    "competencia": "10/2026",
    "registros": [
        {"matricula": "800001", "segurado": "Sistema Parceiro Um",
         "cpf": "123.456.789-00", "capital_morte": 100000,
         "capital_invalidez": 100000, "premio": 40.50,
         "codigo_modulo": "101", "codigo_sub": "01",
         "competencia": "10/2026", "status": "Pago"},
        {"matricula": "800002", "segurado": "Sistema Parceiro Dois",
         "premio": 55.00, "competencia": "10/2026", "status": "A pagar"},
    ],
}
r = api.post("/api/v1/movimentacao", json=lote, headers=CHAVE)
verificar("recebe movimentacao via API (201)", r.status_code == 201, f"veio {r.status_code}")
if r.status_code == 201:
    verificar("  confirma 2 registros gravados", r.json()["registros_gravados"] == 2)

db = SessionLocal()
verificar("  os registros chegaram ao banco",
          db.query(Payment).filter(Payment.competencia == "10/2026").count() == 2)
db.close()

# Reenviar substitui, nao duplica.
api.post("/api/v1/movimentacao", json=lote, headers=CHAVE)
db = SessionLocal()
verificar("  reenviar o lote substitui, nao duplica",
          db.query(Payment).filter(Payment.competencia == "10/2026").count() == 2)
db.close()

# Validacoes de entrada.
ruim = {"competencia": "10/2026", "registros": [
    {"matricula": "1", "segurado": "X", "premio": -5, "competencia": "10/2026"}]}
verificar("premio negativo e recusado (422)",
          api.post("/api/v1/movimentacao", json=ruim, headers=CHAVE).status_code == 422)

divergente = {"competencia": "10/2026", "registros": [
    {"matricula": "1", "segurado": "X", "premio": 10, "competencia": "11/2026"}]}
verificar("competencia divergente e recusada (422)",
          api.post("/api/v1/movimentacao", json=divergente, headers=CHAVE).status_code == 422)

verificar("a documentacao automatica /docs abre",
          TestClient(app).get("/docs").status_code == 200)

# ---------------------------------------------------------------
# 5. LIMPEZA
# ---------------------------------------------------------------
print("\n5. Limpeza")

db = SessionLocal()
n1 = db.query(Payment).filter(Payment.competencia.in_(["09/2026", "10/2026"])).delete()
n2 = db.query(LoginHistory).delete()
db.commit()
restantes = db.query(Payment).count()
db.close()

verificar(f"removidos {n1} pagamentos de teste e {n2} logins", True)
verificar("o banco voltou aos 10 pagamentos originais", restantes == 10,
          f"restaram: {restantes}")

print("\n" + "=" * 50)
print(f"RESULTADO: {passou} passaram, {falhou} falharam")
print("=" * 50 + "\n")

sys.exit(1 if falhou else 0)
