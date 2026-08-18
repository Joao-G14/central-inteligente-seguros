"""
planilha.py
-----------
Le a planilha de movimentacao que a corretora envia e transforma as
linhas dela em registros do banco.

O ARQUIVO ESPERADO
------------------
Mesmo formato da Base_Segurados_Central.xlsx, com estas colunas na
primeira linha (a ordem nao importa, o nome sim):

    NOME DO SEGURADO
    NUMERO DA MATRICULA
    NUMERO DO CPF
    DATA DE NASCIMENTO
    VALOR DO CAPITAL MORTE
    VALOR DO CAPITAL INVALIDEZ
    VALOR DO PREMIO TOTAL/LIQUIDO
    CODIGO DO MODULO
    CODIGO SUB
    COMPETENCIA (dados pagamento)

Uma coluna opcional tambem e aceita:

    SITUACAO DO PAGAMENTO   ->  Pago / A pagar / Em atraso

Se ela nao existir, todos entram como "A pagar".

COMO ESTE ARQUIVO ESTA ORGANIZADO
---------------------------------
1. ler_planilha()  -> abre o arquivo e devolve as linhas + os erros
2. gravar()        -> grava no banco as linhas que passaram
"""

import unicodedata
from datetime import datetime

from openpyxl import load_workbook

from app.models import Payment

# Quantas linhas no maximo aceitamos, para um arquivo gigante nao
# derrubar o servidor.
LIMITE_DE_LINHAS = 5000

# Situacoes de pagamento aceitas.
SITUACOES = ["Pago", "A pagar", "Em atraso"]


def _normalizar(texto) -> str:
    """
    Deixa o nome da coluna comparavel: sem acento, sem espaco sobrando,
    tudo em maiusculo. Assim "Competência" e "COMPETENCIA" viram iguais.
    """
    if texto is None:
        return ""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


# Para cada campo que precisamos, quais nomes de coluna aceitamos.
# Se a corretora mudar levemente o titulo, ainda funciona.
COLUNAS = {
    "segurado": ["NOME DO SEGURADO", "SEGURADO", "NOME"],
    "matricula": ["NUMERO DA MATRICULA", "MATRICULA"],
    "cpf": ["NUMERO DO CPF", "CPF"],
    "nascimento": ["DATA DE NASCIMENTO", "NASCIMENTO"],
    "capital_morte": ["VALOR DO CAPITAL MORTE", "CAPITAL MORTE", "CAP. MORTE"],
    "capital_invalidez": ["VALOR DO CAPITAL INVALIDEZ", "CAPITAL INVALIDEZ", "CAP. INVALIDEZ"],
    "premio": ["VALOR DO PREMIO TOTAL/LIQUIDO", "VALOR DO PREMIO", "PREMIO"],
    "modulo": ["CODIGO DO MODULO", "MODULO"],
    "sub": ["CODIGO SUB", "SUB"],
    "competencia": ["COMPETENCIA (DADOS PAGAMENTO)", "COMPETENCIA"],
    "situacao": ["SITUACAO DO PAGAMENTO", "SITUACAO", "PAGAMENTO"],
}

# Campos sem os quais a linha nao serve para nada.
OBRIGATORIOS = ["segurado", "matricula", "premio", "competencia"]


def _achar_colunas(cabecalho: list) -> dict:
    """
    Descobre em qual posicao esta cada coluna.

    Recebe a primeira linha da planilha e devolve algo como
    {"segurado": 0, "matricula": 1, ...}
    """
    posicoes = {}
    titulos = [_normalizar(c) for c in cabecalho]

    for campo, nomes_aceitos in COLUNAS.items():
        for nome in nomes_aceitos:
            if _normalizar(nome) in titulos:
                posicoes[campo] = titulos.index(_normalizar(nome))
                break

    return posicoes


def _para_numero(valor) -> float:
    """
    Transforma o que veio da celula em numero.

    Aceita 1234.5, "1.234,50" e "R$ 1.234,50".
    Devolve 0.0 se nao conseguir.
    """
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).replace("R$", "").strip()
    # Formato brasileiro: ponto separa milhar, virgula separa decimal.
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _para_texto(valor) -> str:
    """Transforma a celula em texto limpo. Numeros perdem o '.0' do final."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


def _para_competencia(valor) -> str:
    """
    Normaliza a competencia para o formato MM/AAAA.

    Aceita "07/2026", "7/2026" e uma data (o Excel as vezes converte
    sozinho '07/2026' em uma data).
    """
    if isinstance(valor, datetime):
        return valor.strftime("%m/%Y")

    texto = _para_texto(valor)
    if "/" in texto:
        partes = texto.split("/")
        if len(partes) >= 2:
            mes, ano = partes[0].strip(), partes[1].strip()
            if mes.isdigit() and ano.isdigit():
                return f"{int(mes):02d}/{ano}"
    return texto


def ler_planilha(conteudo: bytes) -> tuple[list[dict], list[str]]:
    """
    Abre o arquivo .xlsx e devolve duas listas:

      - as linhas validas, ja convertidas
      - os erros encontrados, em portugues, para mostrar na tela

    Nada e gravado aqui. Ler e gravar sao passos separados de proposito:
    assim, se a planilha tiver problema, o banco nem chega a ser tocado.
    """
    erros: list[str] = []

    try:
        # read_only deixa a leitura rapida e leve;
        # data_only pega o RESULTADO das formulas, nao a formula em si.
        planilha = load_workbook(
            filename=__import__("io").BytesIO(conteudo),
            read_only=True,
            data_only=True,
        )
    except Exception:
        return [], ["Não consegui abrir o arquivo. Ele é mesmo um .xlsx do Excel?"]

    aba = planilha.worksheets[0]
    linhas = list(aba.iter_rows(values_only=True))
    planilha.close()

    if len(linhas) < 2:
        return [], ["A planilha está vazia (só tem o cabeçalho, ou nem isso)."]

    posicoes = _achar_colunas(list(linhas[0]))

    faltando = [c for c in OBRIGATORIOS if c not in posicoes]
    if faltando:
        nomes = ", ".join(COLUNAS[c][0] for c in faltando)
        return [], [f"A planilha não tem estas colunas obrigatórias: {nomes}."]

    def celula(linha, campo):
        indice = posicoes.get(campo)
        if indice is None or indice >= len(linha):
            return None
        return linha[indice]

    registros: list[dict] = []

    # enumerate(..., start=2) porque a linha 1 e o cabecalho: assim o
    # numero do erro bate com o que a pessoa ve no Excel.
    for numero, linha in enumerate(linhas[1:], start=2):
        if numero - 1 > LIMITE_DE_LINHAS:
            erros.append(
                f"A planilha tem mais de {LIMITE_DE_LINHAS} linhas. "
                f"Só as primeiras foram lidas."
            )
            break

        # Pula linhas totalmente vazias e a linha de TOTAL do rodape.
        segurado = _para_texto(celula(linha, "segurado"))
        if not segurado or segurado.upper() == "TOTAL":
            continue

        matricula = _para_texto(celula(linha, "matricula"))
        competencia = _para_competencia(celula(linha, "competencia"))
        premio = _para_numero(celula(linha, "premio"))

        if not matricula:
            erros.append(f"Linha {numero} ({segurado}): sem número de matrícula.")
            continue
        if not competencia:
            erros.append(f"Linha {numero} ({segurado}): sem competência.")
            continue
        if premio <= 0:
            erros.append(f"Linha {numero} ({segurado}): prêmio zerado ou inválido.")
            continue

        situacao = _para_texto(celula(linha, "situacao")) or "A pagar"
        if situacao not in SITUACOES:
            # Nao rejeitamos a linha por causa disso: avisamos e seguimos.
            erros.append(
                f'Linha {numero} ({segurado}): situação "{situacao}" não '
                f'reconhecida, usei "A pagar".'
            )
            situacao = "A pagar"

        registros.append(
            {
                "competencia": competencia,
                "matricula": matricula,
                "segurado": segurado,
                "cpf": _para_texto(celula(linha, "cpf")) or None,
                "capital_morte": _para_numero(celula(linha, "capital_morte")),
                "capital_invalidez": _para_numero(celula(linha, "capital_invalidez")),
                "premio": premio,
                "codigo_modulo": _para_texto(celula(linha, "modulo")) or None,
                "codigo_sub": _para_texto(celula(linha, "sub")) or None,
                "status": situacao,
            }
        )

    if not registros and not erros:
        erros.append("Nenhuma linha com dados foi encontrada na planilha.")

    return registros, erros


def gravar(db, registros: list[dict]) -> dict:
    """
    Grava no banco as linhas lidas.

    REGRA: a planilha SUBSTITUI a competência inteira.
    Se voce enviar a planilha de 08/2026, tudo que existia de 08/2026 e
    apagado e entra o conteudo novo. As outras competencias nao sao
    tocadas. Assim, reenviar a mesma planilha corrigida nao duplica nada.

    Devolve um resumo para mostrar na tela.
    """
    competencias = sorted({r["competencia"] for r in registros})

    apagados = 0
    for competencia in competencias:
        apagados += (
            db.query(Payment).filter(Payment.competencia == competencia).delete()
        )

    db.add_all([Payment(**registro) for registro in registros])
    db.commit()

    return {
        "gravados": len(registros),
        "apagados": apagados,
        "competencias": competencias,
        "premio_total": sum(r["premio"] for r in registros),
    }
